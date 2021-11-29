from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment

from common.views import (
    auto_set_width,
    range_border_internal,
    set_alignment,
    SIDE_THIN
)
from common.models import IngredientCategory
from purchasing.models import ProjectPurchaseOrder, ProjectPurchaseOrderItem


def generate_summary(wb, items):
    sheet_name = '汇总'
    sheet = wb.create_sheet(sheet_name)

    column_category = 1
    column_name = 2
    column_quantity = 3
    column_unit_cost = 4
    column_total_cost = 5

    header_row = 1
    sheet.cell(row=header_row, column=column_category).value = '类别'
    sheet.cell(row=header_row, column=column_name).value = '品名'
    sheet.cell(row=header_row, column=column_quantity).value = '总数量'
    sheet.cell(row=header_row, column=column_unit_cost).value = '单价'
    sheet.cell(row=header_row, column=column_total_cost).value = '总价'
    sheet.row_dimensions[header_row].height = 30

    category_start_row = header_row + 2
    for category in IngredientCategory:
        items_by_category = items.filter(ingredient__category=category)
        ingredient_quantities = defaultdict(float)
        for item in items_by_category:
            ingredient_quantities[item.ingredient] += item.quantity

        num_items = len(ingredient_quantities)
        if not items_by_category:
            sheet.cell(row=category_start_row, column=column_category).value = category.label
            sheet.merge_cells(
                start_row=category_start_row,
                start_column=column_name,
                end_row=category_start_row,
                end_column=column_total_cost
            )
            sheet.cell(row=category_start_row, column=column_name).value = '无'

            range_border_internal(sheet.iter_rows(
                min_row=category_start_row,
                min_col=column_category,
                max_row=category_start_row,
                max_col=column_total_cost
            ), SIDE_THIN, SIDE_THIN)
        else:
            sheet.merge_cells(
                start_row=category_start_row,
                start_column=column_category,
                end_row=category_start_row + num_items - 1,
                end_column=column_category
            )
            sheet.cell(row=category_start_row, column=column_category).value = category.label

            item_row = category_start_row
            print(ingredient_quantities)
            for ingredient, quantity in ingredient_quantities.items():
                sheet.cell(row=item_row, column=column_name).value = ingredient.name
                sheet.cell(row=item_row, column=column_quantity).value = quantity
                sheet.cell(row=item_row, column=column_unit_cost).value = ingredient.price
                total_cost = ingredient.price * quantity
                sheet.cell(row=item_row, column=column_total_cost).value = total_cost
                item_row += 1

            range_border_internal(sheet.iter_rows(
                min_row=category_start_row,
                min_col=column_category,
                max_row=category_start_row + num_items - 1,
                max_col=column_total_cost
            ), SIDE_THIN, SIDE_THIN)

        # There is still a row if num_items == 0
        category_start_row += max(num_items, 1) + 1

    auto_set_width(sheet)
    set_alignment(sheet, Alignment(horizontal='center', vertical="center"))


def generate_project_purchase_order(wb, order, items):
    sheet_name = str(order.project)
    sheet = wb.create_sheet(sheet_name)

    column_category = 1
    column_name = 2
    column_quantity = 3

    header_row = 1
    sheet.cell(row=header_row, column=column_category).value = '类别'
    sheet.cell(row=header_row, column=column_name).value = '品名'
    sheet.cell(row=header_row, column=column_quantity).value = '数量'
    sheet.row_dimensions[header_row].height = 30

    category_start_row = header_row + 2
    for category in IngredientCategory:
        items_by_category = items.filter(ingredient__category=category)
        num_items = len(items_by_category)

        if not items_by_category:
            sheet.cell(row=category_start_row, column=column_category).value = category.label
            sheet.merge_cells(
                start_row=category_start_row,
                start_column=column_name,
                end_row=category_start_row,
                end_column=column_quantity
            )
            sheet.cell(row=category_start_row, column=column_name).value = '无'

            range_border_internal(sheet.iter_rows(
                min_row=category_start_row,
                min_col=column_category,
                max_row=category_start_row,
                max_col=column_quantity
            ), SIDE_THIN, SIDE_THIN)
        else:
            sheet.merge_cells(
                start_row=category_start_row,
                start_column=column_category,
                end_row=category_start_row + num_items - 1,
                end_column=column_category
            )
            sheet.cell(row=category_start_row, column=column_category).value = category.label

            item_row = category_start_row
            for item in items_by_category:
                sheet.cell(row=item_row, column=column_name).value = item.ingredient.name
                sheet.cell(row=item_row, column=column_quantity).value = item.quantity
                item_row += 1

            range_border_internal(sheet.iter_rows(
                min_row=category_start_row,
                min_col=column_category,
                max_row=category_start_row + num_items - 1,
                max_col=column_quantity
            ), SIDE_THIN, SIDE_THIN)

        # There is still a row if num_items == 0
        category_start_row += max(num_items, 1) + 1

    auto_set_width(sheet)
    set_alignment(sheet, Alignment(horizontal='center', vertical="center"))


def generate_project_purchase_orders(wb, date, items):
    orders = ProjectPurchaseOrder.objects.filter(date=date)
    for order in orders:
        project_items = items.filter(order__project=order.project)
        generate_project_purchase_order(wb, order, project_items)


def download_purchase_order_summary(date):
    wb = Workbook()
    del wb['Sheet']

    items = ProjectPurchaseOrderItem.objects.filter(order__date=date)
    generate_summary(wb, items)
    generate_project_purchase_orders(wb, date, items)
    return wb
