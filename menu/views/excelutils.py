from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, BORDER_MEDIUM, BORDER_THIN, Side
from openpyxl.utils import get_column_letter

THIN_BORDER = Border(
    left=Side(style=BORDER_THIN),
    right=Side(style=BORDER_THIN),
    top=Side(style=BORDER_THIN),
    bottom=Side(style=BORDER_THIN)
)

SIDE_THIN = Side(style=BORDER_THIN)
SIDE_MEDIUM = Side(style=BORDER_MEDIUM)


# the ws parameter exists for backward compatability
def range_border_sheet(ws, cell_range, border):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    rows = cell_range
    for cell in rows[0]:
        cell.top = cell.top + top
    for cell in rows[-1]:
        cell.bottom = cell.bottom + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right


def range_border(cell_range, border):
    range_border_sheet(None, list(cell_range), border)


def range_border_internal(cell_range, external, internal):
    """
    I completely have no idea why I have to use += instead of = for border assignment.
    If you change += to =, it will no longer work! Weird!
    :param cell_range: A range or cells
    :param external: The side to be used as the external border
    :param internal: The side to be used as the internal border
    :return: None
    """
    rows = list(cell_range)
    num_rows = len(rows)
    num_columns = len(rows[0])

    # Set the for corners
    rows[0][0].border += Border(top=external, bottom=internal, left=external, right=internal)
    rows[0][-1].border += Border(top=external, bottom=internal, left=internal, right=external)
    rows[-1][0].border += Border(top=internal, bottom=external, left=external, right=internal)
    rows[-1][-1].border += Border(top=internal, bottom=external, left=internal, right=external)

    # Set the four sides
    for i in range(1, num_rows):
        rows[i][0].border += Border(top=internal, bottom=internal, left=external, right=internal)
        rows[i][-1].border += Border(top=internal, bottom=internal, left=internal, right=external)
    for i in range(1, num_columns):
        rows[0][i].border += Border(top=external, bottom=internal, left=internal, right=internal)
        rows[-1][i].border += Border(top=internal, bottom=external, left=internal, right=internal)

    # Set the internal cells
    for i in range(1, num_rows):
        for j in range(1, num_columns):
            internal_border = Border(top=internal, bottom=internal, left=internal, right=internal)
            rows[i][j].border += internal_border


def is_merged_horizontally(ws, cell):
    """
    Checks if cell is merged horizontally with an another cell
    @param cell: cell to check
    @return: True if cell is merged horizontally with an another cell, else False
    """
    cell_coor = cell.coordinate
    if cell_coor not in ws.merged_cells:
        return False
    for rng in ws.merged_cells.ranges:
        if cell_coor in rng and len(list(rng.cols)) > 1:
            return True
    return False


def auto_set_width(ws):
    """
    Adjust width of the columns
    @param ws: worksheet
    @return: None
    """

    def adjusted_len(s):
        # Assume we are mixing ascii and Chinese characters
        return sum(1 if c.isascii() else 2 for c in s)

    for col_number, col in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(col_number)

        max_length = max(
            adjusted_len(str(cell.value or ""))
            for cell in col
            if not is_merged_horizontally(ws, cell)
        )
        adjusted_width = max_length + 0.5
        ws.column_dimensions[col_letter].width = adjusted_width


def set_column_or_row_font(cor, font, exclude=None):
    for cell in cor:
        if exclude and exclude(cell):
            continue
        cell.font = font


def set_column_or_row_color(cor, color, exclude=None):
    for cell in cor:
        if exclude and exclude(cell):
            continue
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')


def set_alignment(ws, alignment):
    for row in ws:
        for cell in row:
            cell.alignment = alignment
