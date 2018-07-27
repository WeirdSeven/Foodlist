import openpyxl

THIN_BORDER = openpyxl.styles.borders.Border(left = openpyxl.styles.borders.Side(style='thin'), 
											 right = openpyxl.styles.borders.Side(style='thin'), 
											 top = openpyxl.styles.borders.Side(style='thin'), 
											 bottom = openpyxl.styles.borders.Side(style='thin'))

def range_border(ws, cell_range, border):
	top = openpyxl.styles.borders.Border(top = border.top)
	left = openpyxl.styles.borders.Border(left = border.left)
	right = openpyxl.styles.borders.Border(right = border.right)
	bottom = openpyxl.styles.borders.Border(bottom = border.bottom)

	rows = cell_range
	for cell in rows[0]:
		cell.border = cell.border + top
	for cell in rows[-1]:
		cell.border = cell.border + bottom

	for row in rows:
		l = row[0]
		r = row[-1]
		l.border = l.border + left
		r.border = r.border + right

def generate_program_ingredient_report(wb, title, program_dishes):
	COLUMN_ID = 1
	COLUMN_DISH = 2
	COLUMN_COUNT = 3
	COLUMN_INGREDIENT = 4
	COLUMN_WEIGHT = 5

	wb.create_sheet('配料单')
	sheet = wb['配料单']

	sheet.merge_cells(start_row = 1, start_column = COLUMN_ID, end_row = 1, end_column = COLUMN_WEIGHT)
	sheet.cell(row = 1, column = 1).value = title
	range_border(sheet, list(sheet.iter_rows(min_row = 1, min_col = COLUMN_ID, max_row = 1, max_col = COLUMN_WEIGHT)), THIN_BORDER)

	sheet.cell(row = 2, column = COLUMN_ID).value = '序号'
	sheet.cell(row = 2, column = COLUMN_DISH).value = '菜品'
	sheet.cell(row = 2, column = COLUMN_COUNT).value = '份数'
	sheet.cell(row = 2, column = COLUMN_INGREDIENT).value = '配料'
	sheet.cell(row = 2, column = COLUMN_WEIGHT).value = '重量'

	sheet.cell(row = 2, column = COLUMN_ID).border = THIN_BORDER
	sheet.cell(row = 2, column = COLUMN_DISH).border = THIN_BORDER
	sheet.cell(row = 2, column = COLUMN_COUNT).border = THIN_BORDER
	sheet.cell(row = 2, column = COLUMN_INGREDIENT).border = THIN_BORDER
	sheet.cell(row = 2, column = COLUMN_WEIGHT).border = THIN_BORDER

	cur_row = 3
	for dish in program_dishes:
		num_ingredients = len(dish[3])
		for col in [COLUMN_ID, COLUMN_DISH, COLUMN_COUNT]:
			sheet.merge_cells(start_row = cur_row, start_column = col, end_row = cur_row + num_ingredients - 1, end_column = col)
			sheet.cell(row = cur_row, column = col).value = dish[col - 1]
			range_border(sheet, list(sheet.iter_rows(min_row = cur_row, min_col = col, max_row = cur_row + num_ingredients - 1, max_col = col)), THIN_BORDER)

		if num_ingredients == 0:
			sheet.merge_cells(start_row = cur_row, start_column = COLUMN_INGREDIENT, end_row = cur_row, end_column = COLUMN_WEIGHT)
			sheet.cell(row = cur_row, column = COLUMN_INGREDIENT).value = '无'
			range_border(sheet, list(shee.iter_rows(min_row = cur_row, min_col = COLUMN_INGREDIENT, max_row = cur_row, max_col = COLUMN_WEIGHT)), THIN_BORDER)
			cur_row += 1
		else:	
			for ingredients in dish[3]:
				sheet.cell(row = cur_row, column = COLUMN_INGREDIENT).value = ingredients[0]
				sheet.cell(row = cur_row, column = COLUMN_WEIGHT).value = ingredients[1]
				sheet.cell(row = cur_row, column = COLUMN_INGREDIENT).border = THIN_BORDER
				sheet.cell(row = cur_row, column = COLUMN_WEIGHT).border = THIN_BORDER
				cur_row += 1

def generate_program_purchase_report(wb, title, program_ingredients):
	COLUMN_ID = 1
	COLUMN_INGREDIENT = 2
	COLUMN_QUANTITY = 3
	COLUMN_COST = 4

	wb.create_sheet('采购清单')
	sheet = wb['采购清单']

	sheet.merge_cells(start_row = 1, start_column = COLUMN_ID, end_row = 1, end_column = COLUMN_COST)
	sheet.cell(row = 1, column = 1).value = title
	range_border(sheet, list(sheet.iter_rows(min_row = 1, min_col = COLUMN_ID, max_row = 1, max_col = COLUMN_COST)), THIN_BORDER)

	sheet.cell(row = 2, column = COLUMN_ID).value = '序号'
	sheet.cell(row = 2, column = COLUMN_INGREDIENT).value = '配料'
	sheet.cell(row = 2, column = COLUMN_QUANTITY).value = '重量'
	sheet.cell(row = 2, column = COLUMN_COST).value = '成本'

	for col in [COLUMN_ID, COLUMN_INGREDIENT, COLUMN_QUANTITY, COLUMN_COST]:
		sheet.cell(row = 2, column = col).border = THIN_BORDER

	cur_row = 3
	for ingredient in program_ingredients:
		for col in [COLUMN_ID, COLUMN_INGREDIENT, COLUMN_QUANTITY, COLUMN_COST]:
			sheet.cell(row = cur_row, column = col).value = ingredient[col - 1]
			sheet.cell(row = cur_row, column = col).border = THIN_BORDER
		cur_row += 1

def program_ingredient_report(title, program_dishes, program_ingredients):
	wb = openpyxl.Workbook()
	generate_program_ingredient_report(wb, title + '配料单', program_dishes)
	generate_program_purchase_report(wb, title + '采购清单', program_ingredients)
	del wb['Sheet']
	return wb


def generate_company_ingredient_report(wb, title, program_names, company_dishes):
	COLUMN_ID = 1
	COLUMN_DISH = 2
	COLUMN_COUNT_START = 3
	COLUMN_COUNT_END = 3 + len(program_names) - 1
	COLUMN_INGREDIENT = COLUMN_COUNT_END + 1
	COLUMN_WEIGHT = COLUMN_INGREDIENT + 1

	wb.create_sheet('配料单')
	sheet = wb['配料单']

	sheet.merge_cells(start_row = 1, start_column = COLUMN_ID, end_row = 1, end_column = COLUMN_WEIGHT)
	sheet.cell(row = 1, column = 1).value = title
	range_border(sheet, list(sheet.iter_rows(min_row = 1, min_col = COLUMN_ID, max_row = 1, max_col = COLUMN_WEIGHT)), THIN_BORDER)

	sheet.merge_cells(start_row = 2, start_column = COLUMN_ID, end_row = 3, end_column = COLUMN_ID)
	sheet.cell(row = 2, column = COLUMN_ID).value = '序号'
	sheet.merge_cells(start_row = 2, start_column = COLUMN_DISH, end_row = 3, end_column = COLUMN_DISH)
	sheet.cell(row = 2, column = COLUMN_DISH).value = '菜品'
	sheet.merge_cells(start_row = 2, start_column = COLUMN_COUNT_START, end_row = 2, end_column = COLUMN_COUNT_END)
	sheet.cell(row = 2, column = COLUMN_COUNT_START).value = '份数'
	for col in range(COLUMN_COUNT_START, COLUMN_COUNT_END + 1):
		sheet.cell(row = 3, column = col).value = program_names[col - COLUMN_COUNT_START]
	sheet.merge_cells(start_row = 2, start_column = COLUMN_INGREDIENT, end_row = 3, end_column = COLUMN_INGREDIENT)
	sheet.cell(row = 2, column = COLUMN_INGREDIENT).value = '配料'
	sheet.merge_cells(start_row = 2, start_column = COLUMN_WEIGHT, end_row = 3, end_column = COLUMN_WEIGHT)
	sheet.cell(row = 2, column = COLUMN_WEIGHT).value = '重量'


	range_border(sheet, list(sheet.iter_rows(min_row = 2, min_col = COLUMN_ID, max_row = 3, max_col = COLUMN_ID)), THIN_BORDER)
	range_border(sheet, list(sheet.iter_rows(min_row = 2, min_col = COLUMN_DISH, max_row = 3, max_col = COLUMN_DISH)), THIN_BORDER)
	range_border(sheet, list(sheet.iter_rows(min_row = 2, min_col = COLUMN_COUNT_START, max_row = 2, max_col = COLUMN_COUNT_END)), THIN_BORDER)
	for col in range(COLUMN_COUNT_START, COLUMN_COUNT_END + 1):
		sheet.cell(row = 3, column = col).border = THIN_BORDER
	range_border(sheet, list(sheet.iter_rows(min_row = 2, min_col = COLUMN_INGREDIENT, max_row = 3, max_col = COLUMN_INGREDIENT)), THIN_BORDER)
	range_border(sheet, list(sheet.iter_rows(min_row = 2, min_col = COLUMN_WEIGHT, max_row = 3, max_col = COLUMN_WEIGHT)), THIN_BORDER)

	
	cur_row = 4
	for dish in company_dishes:
		num_ingredients = len(dish[3])
		for col in [COLUMN_ID, COLUMN_DISH]:
			sheet.merge_cells(start_row = cur_row, start_column = col, end_row = cur_row + num_ingredients - 1, end_column = col)
			sheet.cell(row = cur_row, column = col).value = dish[col - 1]
			range_border(sheet, list(sheet.iter_rows(min_row = cur_row, min_col = col, max_row = cur_row + num_ingredients - 1, max_col = col)), THIN_BORDER)

		for i, count in enumerate(dish[2]):
			sheet.merge_cells(start_row = cur_row, start_column = COLUMN_COUNT_START + i, end_row = cur_row + num_ingredients - 1, end_column = COLUMN_COUNT_START + i)
			sheet.cell(row = cur_row, column = COLUMN_COUNT_START + i).value = count
			range_border(sheet, list(sheet.iter_rows(min_row = cur_row, min_col = COLUMN_COUNT_START + i, max_row = cur_row + num_ingredients - 1, max_col = COLUMN_COUNT_START + i)), THIN_BORDER)

		if num_ingredients == 0:
			sheet.merge_cells(start_row = cur_row, start_column = COLUMN_INGREDIENT, end_row = cur_row, end_column = COLUMN_WEIGHT)
			sheet.cell(row = cur_row, column = COLUMN_INGREDIENT).value = '无'
			range_border(sheet, list(shee.iter_rows(min_row = cur_row, min_col = COLUMN_INGREDIENT, max_row = cur_row, max_col = COLUMN_WEIGHT)), THIN_BORDER)
			cur_row += 1
		else:	
			for ingredients in dish[3]:
				sheet.cell(row = cur_row, column = COLUMN_INGREDIENT).value = ingredients[0]
				sheet.cell(row = cur_row, column = COLUMN_WEIGHT).value = ingredients[1]
				sheet.cell(row = cur_row, column = COLUMN_INGREDIENT).border = THIN_BORDER
				sheet.cell(row = cur_row, column = COLUMN_WEIGHT).border = THIN_BORDER
				cur_row += 1
		

def generate_company_purchase_report(wb, title, company_ingredients):
	generate_program_purchase_report(wb, title, company_ingredients)

def company_ingredient_report(title, program_names, company_dishes, company_ingredients):
	wb = openpyxl.Workbook()
	generate_company_ingredient_report(wb, title + '公司总配料单', program_names, company_dishes)
	generate_company_purchase_report(wb, title + '公司总采购清单', company_ingredients)
	del wb['Sheet']
	return wb




