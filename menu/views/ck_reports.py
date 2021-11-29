# -*- coding: utf-8 -*-

from collections import defaultdict
from functools import partial

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from common.models import SDish2StandardIngredient
from common.views import (
    auto_set_width,
    is_merged_horizontally,
    range_border_internal,
    set_alignment,
    set_column_or_row_font,
    set_column_or_row_color,
    SIDE_MEDIUM,
    SIDE_THIN
)
from menu.models import (
    CKProject2SDish2Standard,
    CKProject2SDish2StandardCount,
    Course,
    Meal
)

weekday_to_name = {1: '周一',
                   2: '周二',
                   3: '周三',
                   4: '周四',
                   5: '周五',
                   6: '周六',
                   7: '周日'}


def generate_ckproject_weekly_report_menu(wb, projects):
    sheet_name = '周菜单'
    wb.create_sheet(sheet_name)
    sheet = wb[sheet_name]

    def merge_max(dict1, dict2):
        all_keys = dict1.keys() | dict2.keys()
        result = {}
        for k in all_keys:
            if k in dict1 and k in dict2:
                result[k] = max(dict1[k], dict2[k])
            elif k in dict1:
                result[k] = dict1[k]
            else:
                result[k] = dict2[k]
        return result

    breakfast_courses = [
        Course.SPECIAL,
        Course.STAPLE,
        Course.SIDE,
        Course.COLD,
        Course.PICKLE,
        Course.SOUP
    ]
    regular_courses = [
        Course.PRIMARY_MEAT,
        Course.SECONDARY_MEAT,
        Course.VEGETABLE,
        Course.SPECIAL,
        Course.FRUIT,
        Course.GRAIN,
        Course.SOUP
    ]
    course_by_meal = {
        Meal.BREAKFAST: breakfast_courses,
        Meal.LUNCH: regular_courses,
        Meal.DINNER: regular_courses,
        Meal.MIDNIGHT: regular_courses
    }

    column_meal = 1
    column_course = 2

    weekdays = [project.date.isoweekday() for project in projects]
    p2ds_week = [CKProject2SDish2Standard.objects.filter(project=project) for project in projects]

    meal_start_row = 1
    for meal in Meal:
        p2ds_meal = [p2ds_day.filter(meal=meal) for p2ds_day in p2ds_week]

        cur_meal_courses = course_by_meal[meal]
        course_counts = {course: 1 for course in cur_meal_courses}
        for p2ds_day in p2ds_meal:
            course_counts_day = defaultdict(int)
            for p2d in p2ds_day:
                course_counts_day[Course(p2d.course)] += 1
            course_counts = merge_max(course_counts, course_counts_day)

        distinct_courses = course_counts.keys()
        extra_courses = distinct_courses - cur_meal_courses
        extra_courses_ordered = [course for course in Course if course in extra_courses]
        all_courses_ordered = cur_meal_courses + extra_courses_ordered

        # The Meal column
        num_rows_meal = sum(course_counts.values())
        sheet.merge_cells(
            start_row=meal_start_row,
            start_column=column_meal,
            end_row=meal_start_row + num_rows_meal,
            end_column=column_meal
        )
        sheet.cell(row=meal_start_row, column=column_meal).value = meal.label + '菜单'

        # The Course column
        sheet.cell(row=meal_start_row, column=column_course).value = '星期项目'
        cur_course_row = meal_start_row + 1
        for course in all_courses_ordered:
            cur_course_count = course_counts[course]
            sheet.merge_cells(
                start_row=cur_course_row,
                start_column=column_course,
                end_row=cur_course_row + cur_course_count - 1,
                end_column=column_course
            )
            sheet.cell(row=cur_course_row, column=column_course).value = course.label
            cur_course_row += cur_course_count

        # The dish columns of weekdays
        for i in weekdays:
            cur_weekday_column = column_course + i
            sheet.cell(row=meal_start_row, column=cur_weekday_column).value = weekday_to_name[i]

            cur_dish_row = meal_start_row + 1
            for course in all_courses_ordered:
                cur_course_count = course_counts[course]
                p2ds_course = p2ds_meal[i - 1].filter(course=course)
                course_dish_names = [str(p2d.sdish2standard) for p2d in p2ds_course]
                course_dish_names += ['无'] * (cur_course_count - len(course_dish_names))

                for dish_name in course_dish_names:
                    sheet.cell(row=cur_dish_row, column=cur_weekday_column).value = dish_name
                    cur_dish_row += 1

        # Set border
        range_border_internal(sheet.iter_rows(
            min_row=meal_start_row,
            min_col=column_meal,
            max_row=meal_start_row + num_rows_meal,
            max_col=column_course + len(weekdays)
        ), SIDE_MEDIUM, SIDE_THIN)

        meal_start_row += num_rows_meal + 2  # 2 comes from a header row and and a blank row

    # Set column widths and cell alignment
    auto_set_width(sheet)
    set_alignment(sheet, Alignment(horizontal='center', vertical="center"))


def get_project_locations_by_meal(p2ds, meal):
    locations = set()
    for p2d in p2ds.filter(meal=meal):
        locations.update(p2d.locations_counts.values_list('location__name', flat=True))
    return list(locations)


def generate_ckproject_weekly_report_day(wb, project):
    weekday = project.date.isoweekday()
    weekday_name = weekday_to_name[weekday]
    wb.create_sheet(weekday_name)
    sheet = wb[weekday_name]

    p2ds = CKProject2SDish2Standard.objects.filter(project=project)
    num_project_locations = max([len(get_project_locations_by_meal(p2ds, meal)) for meal in Meal])
    num_ingredients = max([p2d.sdish2standard.ingredients.count() for p2d in p2ds])
    num_ingredient_groups = (num_ingredients + 1) // 2  # Two ingredients per group
    num_columns_per_ingredient_group = 3  # Name, quantity, total

    column_dish_name = 1
    column_first_location = 2
    column_total = column_dish_name + num_project_locations + 1
    column_first_ingredient = column_total + 1
    column_last = column_total + num_columns_per_ingredient_group * num_ingredient_groups

    total_num_columns = (
        1  # Name of dish
        + num_project_locations
        + 1  # Total
        + num_ingredient_groups * num_columns_per_ingredient_group
    )

    default_row_height = 15

    meal_start_row = 1
    for meal in Meal:
        sheet.merge_cells(
            start_row=meal_start_row,
            start_column=column_dish_name,
            end_row=meal_start_row,
            end_column=total_num_columns
        )
        title_cell = sheet.cell(row=meal_start_row, column=1)
        title_cell.value = '%s %s %s %d月%d日' % (
            project.name,
            weekday_name,
            meal.label,
            project.date.month,
            project.date.day
        )
        title_cell.font = Font(bold=True)
        sheet.row_dimensions[meal_start_row].height = 30

        # Fill in the header row
        header_row = meal_start_row + 1
        sheet.cell(row=header_row, column=column_dish_name).value = '菜品名称'
        locations = get_project_locations_by_meal(p2ds, meal)
        for col, location in enumerate(locations, start=column_first_location):
            sheet.cell(row=header_row, column=col).value = location
        sheet.cell(row=header_row, column=column_total).value = '合计'
        for i in range(num_ingredient_groups):
            sheet.cell(
                row=header_row,
                column=column_first_ingredient + i * num_columns_per_ingredient_group
            ).value = '原材料'
            sheet.cell(
                row=header_row,
                column=column_first_ingredient + i * num_columns_per_ingredient_group + 1
            ).value = '单人量（斤）'
            sheet.cell(
                row=header_row,
                column=column_first_ingredient + i * num_columns_per_ingredient_group + 2
            ).value = '采购量（斤）'
        set_column_or_row_font(sheet[header_row], Font(bold=True))
        sheet.row_dimensions[header_row].height = default_row_height * 2

        # Fill in the dish rows
        cur_dish_row = header_row + 1
        p2ds_of_meal = p2ds.filter(meal=meal)
        if not p2ds_of_meal:
            sheet.merge_cells(
                start_row=cur_dish_row,
                start_column=column_dish_name,
                end_row=cur_dish_row,
                end_column=column_last
            )
            sheet.cell(row=cur_dish_row, column=column_dish_name).value = '无餐'
            cur_dish_row += 1
        else:
            for p2d in p2ds_of_meal:
                for i in range(column_dish_name, column_total + 1):
                    sheet.merge_cells(
                        start_row=cur_dish_row,
                        start_column=i,
                        end_row=cur_dish_row + 1,
                        end_column=i
                    )

                # Dish name
                sheet.cell(
                    row=cur_dish_row,
                    column=column_dish_name
                ).value = str(p2d.sdish2standard)

                # Delivery locations and counts
                p2dc_all = CKProject2SDish2StandardCount.objects.filter(project2dish2standard=p2d)
                total_count = 0
                for col, location in enumerate(locations, start=column_first_location):
                    counts = p2dc_all.filter(location__name=location)
                    count = counts[0].count if len(counts) == 1 else 0
                    total_count += count
                    sheet.cell(row=cur_dish_row, column=col).value = count if count != 0 else ''
                sheet.cell(row=cur_dish_row, column=column_total).value = total_count

                # Ingredients and quantities
                cur_ingredient_row = cur_dish_row
                cur_ingredient_column = column_first_ingredient
                s2s_ingredients = SDish2StandardIngredient.objects.filter(
                    sdish2standard=p2d.sdish2standard
                )
                for i, s2s_ingredient in enumerate(s2s_ingredients):
                    ingredient = s2s_ingredient.ingredient
                    quantity = s2s_ingredient.quantity

                    sheet.cell(
                        row=cur_ingredient_row,
                        column=cur_ingredient_column
                    ).value = str(ingredient)
                    sheet.cell(
                        row=cur_ingredient_row,
                        column=cur_ingredient_column + 1
                    ).value = quantity
                    sheet.cell(
                        row=cur_ingredient_row,
                        column=cur_ingredient_column + 2
                    ).value = quantity * total_count

                    if i % 2 == 0:
                        cur_ingredient_row += 1
                    else:
                        cur_ingredient_row -= 1
                        cur_ingredient_column += 3

                # A dish spans two rows
                cur_dish_row += 2

        # Set border
        range_border_internal(sheet.iter_rows(
            min_row=header_row,
            min_col=column_dish_name,
            max_row=cur_dish_row - 1,
            max_col=column_last
        ), SIDE_MEDIUM, SIDE_THIN)

        meal_start_row = cur_dish_row

    # Set color for certain columns
    columns = [None] + list(sheet.columns)  # column indices start from 1
    yellow_rgb = 'FFFF00'
    set_column_or_row_color(
        columns[column_dish_name],
        yellow_rgb,
        exclude=partial(is_merged_horizontally, sheet)
    )
    set_column_or_row_color(columns[column_total], yellow_rgb)
    for i in range(num_ingredient_groups):
        column_purchase_quantity = column_total + i * num_columns_per_ingredient_group
        set_column_or_row_color(columns[column_purchase_quantity], yellow_rgb)

    # Set column widths and cell alignment
    auto_set_width(sheet)
    set_alignment(sheet, Alignment(horizontal='center', vertical="center"))


def generate_ckproject_weekly_report(projects):
    wb = Workbook()
    del wb['Sheet']

    generate_ckproject_weekly_report_menu(wb, projects)

    for project in projects:
        generate_ckproject_weekly_report_day(wb, project)

    return wb
